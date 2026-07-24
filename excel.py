import csv
import io
from typing import List, Dict, Any, Union, BinaryIO, Tuple
from config import Config
from logger import logger

# Try importing pandas; if Windows Application Control blocks pandas C-extensions, fallback gracefully to openpyxl & csv
HAS_PANDAS = False
try:
    import pandas as pd
    HAS_PANDAS = True
except Exception as e:
    logger.warning(f"Pandas import unavailable ({e}). Using openpyxl & csv fallbacks.")
    pd = None

try:
    import openpyxl
except ImportError:
    openpyxl = None


def read_input_file(file_input: Union[str, BinaryIO]) -> Tuple[List[Dict[str, Any]], List[str]]:
    """Reads Excel (.xlsx) or CSV files into a list of row dictionaries and column header list."""
    records = []
    headers = []

    try:
        if HAS_PANDAS:
            if isinstance(file_input, str):
                df = pd.read_csv(file_input) if file_input.endswith(".csv") else pd.read_excel(file_input)
            else:
                filename = getattr(file_input, "name", "").lower()
                df = pd.read_csv(file_input) if filename.endswith(".csv") else pd.read_excel(file_input)
            
            df = df.fillna("")
            headers = list(df.columns)
            records = df.to_dict(orient="records")
            return records, headers

        # Fallback reading without pandas
        if isinstance(file_input, str):
            if file_input.endswith(".csv"):
                with open(file_input, mode="r", encoding="utf-8-sig") as f:
                    reader = csv.DictReader(f)
                    headers = reader.fieldnames or []
                    records = [dict(row) for row in reader]
            else:
                wb = openpyxl.load_workbook(file_input, data_only=True)
                sheet = wb.active
                rows = list(sheet.iter_rows(values_only=True))
                if rows:
                    headers = [str(cell or "").strip() for cell in rows[0]]
                    for row in rows[1:]:
                        records.append({headers[i]: (row[i] if i < len(row) and row[i] is not None else "") for i in range(len(headers))})
        else:
            filename = getattr(file_input, "name", "").lower()
            if filename.endswith(".csv"):
                content = file_input.read().decode("utf-8-sig")
                reader = csv.DictReader(io.StringIO(content))
                headers = reader.fieldnames or []
                records = [dict(row) for row in reader]
            else:
                wb = openpyxl.load_workbook(file_input, data_only=True)
                sheet = wb.active
                rows = list(sheet.iter_rows(values_only=True))
                if rows:
                    headers = [str(cell or "").strip() for cell in rows[0]]
                    for row in rows[1:]:
                        records.append({headers[i]: (row[i] if i < len(row) and row[i] is not None else "") for i in range(len(headers))})

        logger.info(f"Loaded file with {len(records)} rows using fallback parser.")
        return records, headers

    except Exception as e:
        logger.error(f"Error reading input file: {e}")
        raise ValueError(f"Could not parse file: {e}")


def merge_classification_results(
    original_records: List[Dict[str, Any]], results: List[Dict[str, Any]]
) -> List[Dict[str, Any]]:
    """Merges structured classification results into record dictionaries."""
    merged = []
    for orig, res in zip(original_records, results):
        row = dict(orig)
        row["Category"] = res.get("category", "")
        row["Urgency"] = res.get("urgency", "")
        row["Sentiment"] = res.get("sentiment", "")
        row["Summary"] = res.get("summary", "")
        row["Suggested Reply"] = res.get("suggested_reply", "")
        key_ents = res.get("key_entities", [])
        row["Key Entities"] = ", ".join(key_ents) if isinstance(key_ents, list) else str(key_ents)
        merged.append(row)
    return merged


def export_to_excel(records: List[Dict[str, Any]], filename: str = "classified_emails.xlsx") -> str:
    """Exports records to an Excel spreadsheet in the output directory."""
    output_path = Config.OUTPUT_DIR / filename

    if not records:
        return str(output_path)

    fieldnames = list(records[0].keys())

    if openpyxl:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Classified Emails"
        ws.append(fieldnames)
        for row in records:
            ws.append([str(row.get(f, "")) for f in fieldnames])
        wb.save(output_path)
    else:
        # Fallback to CSV if openpyxl is absent
        csv_path = Config.OUTPUT_DIR / filename.replace(".xlsx", ".csv")
        with open(csv_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(records)
        return str(csv_path)

    logger.info(f"Successfully generated export at: {output_path}")
    return str(output_path)


def create_sample_excel(filename: str = "sample_customer_emails.xlsx") -> str:
    """Generates a sample customer emails dataset for testing."""
    sample_data = [
        {
            "Email_ID": "MSG-1006",
            "Subject": "Webhook events dropping during high traffic periods",
            "Body": "Hello support, we noticed that webhook delivery events are failing with 429 status codes between 2 PM and 4 PM UTC. Is there a rate limit boost available for our account?",
        },
        {
            "Email_ID": "MSG-1007",
            "Subject": "Switching from monthly to annual billing plan",
            "Body": "Hi, our finance team would like to transition our current Pro subscription to an annual plan. Could you apply the 15% annual discount and send the updated quote?",
        },
        {
            "Email_ID": "MSG-1008",
            "Subject": "Inquiry regarding SOC 2 Type II compliance and data residency",
            "Body": "Dear Sales Team, We are looking to adopt your product for our healthcare organization. Before proceeding, can you confirm if customer data can be hosted exclusively within the EU region and share your latest SOC 2 report?",
        },
        {
            "Email_ID": "MSG-1009",
            "Subject": "Feedback: Dark mode contrast issues on mobile web browser",
            "Body": "Hey team, the dark mode toggle on mobile browsers makes text in dropdown menus hard to read due to low contrast. Adding an accessibility fix would be awesome!",
        },
        {
            "Email_ID": "MSG-1010",
            "Subject": "URGENT: Suspicious login notification from unrecognized IP",
            "Body": "I received an automated email alerting me of a login from an IP in another country. I have locked my account, but need assistance resetting 2FA hardware keys immediately.",
        },
    ]

    output_path = Config.DATA_DIR / filename
    if openpyxl:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sample Emails"
        ws.append(list(sample_data[0].keys()))
        for r in sample_data:
            ws.append(list(r.values()))
        wb.save(output_path)
    return str(output_path)

